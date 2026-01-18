"""
Check Excel file headers and compare with expected headers
"""
import os
import sys

try:
    import openpyxl
except ImportError:
    print("Installing openpyxl...")
    os.system(f"{sys.executable} -m pip install openpyxl")
    import openpyxl

# Expected headers from the code
EXPECTED_HEADERS = [
    'Item name*',
    'Item code',
    'Description',
    'Category',
    'Base Unit (x)',
    'Secondary Unit (y)',
    'Conversion Rate (n) (x = ny)',
    'Supplier',
    'markup_Margin',
    'Price_List_Retail_Price',
    'Price_List_Wholesale_Price',
    'Price_List_Trade_Price',
    'Price_List_Last_Cost',
    'Price_List_Average_Cost',
    'Price_List_Retail_Unit_Price',
    'Price_List_Wholesale_Unit_Price',
    'Price_List_Trade_Unit_Price',
    'Price_List_Tax_Code',
    'Price_List_Tax_Percentage',
    'Price_List_Tax_Description',
    'Price_List_Tax_Type',
    'Price_List_Price_Inclusive',
    'Current stock quantity',
    'Minimum stock quantity',
    'HSN',
    'Sale Discount',
    'Tax Rate',
    'Inclusive Of Tax',
    'Price_List_Min_Price',
    'Price_List_Special_Price',
    'Price_List_Has_Refill',
    'Price_List_Not_For_Sale',
    'Price_List_Is_Physical_Item',
    'Price_List_Min_Wholesale_Price',
    'Price_List_Min_Wholesale_Unit_Price',
    'Price_List_Min_Retail_Price',
    'Price_List_Min_Retail_Unit_Price',
    'Price_List_Min_Trade_Price',
    'Price_List_Min_Trade_Unit_Price'
]

excel_path = r"C:\Users\Envy\Downloads\pharmasight_template.xlsx"

print("=" * 80)
print("Excel File Header Analysis")
print("=" * 80)
print(f"\nFile: {excel_path}")
print(f"File exists: {os.path.exists(excel_path)}")

if not os.path.exists(excel_path):
    print("\nERROR: File not found!")
    sys.exit(1)

try:
    # Load workbook
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    sheet = wb.active
    
    print(f"\nSheet name: {sheet.title}")
    print(f"Total rows: {sheet.max_row}")
    print(f"Total columns: {sheet.max_column}")
    
    # Get headers from first row
    actual_headers = []
    for cell in sheet[1]:
        header = str(cell.value).strip() if cell.value else ""
        actual_headers.append(header)
    
    print(f"\nActual headers found ({len(actual_headers)}):")
    print("-" * 80)
    for i, header in enumerate(actual_headers, 1):
        print(f"{i:2}. {header}")
    
    print(f"\nExpected headers ({len(EXPECTED_HEADERS)}):")
    print("-" * 80)
    for i, header in enumerate(EXPECTED_HEADERS, 1):
        print(f"{i:2}. {header}")
    
    # Compare
    print("\n" + "=" * 80)
    print("COMPARISON")
    print("=" * 80)
    
    actual_set = set([h.lower().strip() for h in actual_headers if h])
    expected_set = set([h.lower().strip() for h in EXPECTED_HEADERS])
    
    missing_in_actual = expected_set - actual_set
    extra_in_actual = actual_set - expected_set
    
    if missing_in_actual:
        print(f"\n❌ Missing in Excel file ({len(missing_in_actual)}):")
        for header in sorted(missing_in_actual):
            # Find original case
            orig = next((h for h in EXPECTED_HEADERS if h.lower().strip() == header), header)
            print(f"   - {orig}")
    
    if extra_in_actual:
        print(f"\n⚠️  Extra in Excel file ({len(extra_in_actual)}):")
        for header in sorted(extra_in_actual):
            # Find original case
            orig = next((h for h in actual_headers if h.lower().strip() == header), header)
            print(f"   - {orig}")
    
    if not missing_in_actual and not extra_in_actual:
        print("\n[OK] Headers match perfectly!")
    else:
        print(f"\n[WARNING] Headers don't match exactly!")
        print(f"   Matched: {len(expected_set & actual_set)}")
        print(f"   Missing: {len(missing_in_actual)}")
        print(f"   Extra: {len(extra_in_actual)}")
    
    # Check critical headers
    print("\n" + "=" * 80)
    print("CRITICAL HEADERS CHECK")
    print("=" * 80)
    
    critical_headers = {
        'Item name*': ['item name*', 'item name', 'name'],
        'Base Unit (x)': ['base unit (x)', 'base unit', 'base unit (x)', 'unit'],
        'Price_List_Last_Cost': ['price_list_last_cost', 'last cost', 'purchase price', 'cost'],
        'Price_List_Tax_Percentage': ['price_list_tax_percentage', 'tax rate', 'tax percentage', 'vat rate'],
    }
    
    for expected, variations in critical_headers.items():
        found = False
        found_header = None
        for actual in actual_headers:
            if actual.lower().strip() in [v.lower() for v in variations]:
                found = True
                found_header = actual
                break
        
        if found:
            print(f"[OK] {expected}: Found as '{found_header}'")
        else:
            print(f"[ERROR] {expected}: NOT FOUND")
    
    # Show sample data and check for issues
    print("\n" + "=" * 80)
    print("SAMPLE DATA ANALYSIS (First 10 rows)")
    print("=" * 80)
    
    issues_found = {
        'empty_name': [],
        'empty_base_unit': [],
        'invalid_cost': [],
        'invalid_vat_rate': [],
    }
    
    for row_idx in range(2, min(12, sheet.max_row + 1)):
        row_data = {}
        for col_idx, header in enumerate(actual_headers, 1):
            cell_value = sheet.cell(row=row_idx, column=col_idx).value
            row_data[header] = cell_value
        
        # Check for issues
        item_name = str(row_data.get('Item name*', '')).strip()
        base_unit = str(row_data.get('Base Unit (x)', '')).strip()
        cost = row_data.get('Price_List_Last_Cost') or row_data.get('Price_List_Average_Cost')
        vat_rate = row_data.get('Price_List_Tax_Percentage') or row_data.get('Tax Rate')
        
        if not item_name or item_name == '' or item_name == 'None':
            issues_found['empty_name'].append(row_idx)
        
        if not base_unit or base_unit == '' or base_unit == 'None':
            issues_found['empty_base_unit'].append(row_idx)
        
        try:
            cost_val = float(cost) if cost else 0
            if cost_val < 0:
                issues_found['invalid_cost'].append((row_idx, cost_val))
        except (ValueError, TypeError):
            if cost and str(cost).strip() not in ['', 'None', '0']:
                issues_found['invalid_cost'].append((row_idx, cost))
        
        try:
            vat_val = float(vat_rate) if vat_rate else 0
            if vat_val < 0 or vat_val > 100:
                issues_found['invalid_vat_rate'].append((row_idx, vat_val))
        except (ValueError, TypeError):
            if vat_rate and str(vat_rate).strip() not in ['', 'None', '0']:
                issues_found['invalid_vat_rate'].append((row_idx, vat_rate))
        
        # Show first 3 rows in detail
        if row_idx <= 4:
            print(f"\nRow {row_idx}:")
            print(f"  Item name*: {item_name}")
            print(f"  Base Unit (x): {base_unit}")
            print(f"  Price_List_Last_Cost: {cost}")
            print(f"  Price_List_Tax_Percentage: {vat_rate}")
            print(f"  Category: {row_data.get('Category', '')}")
    
    # Summary of issues
    print("\n" + "=" * 80)
    print("DATA QUALITY ISSUES FOUND")
    print("=" * 80)
    
    total_issues = 0
    if issues_found['empty_name']:
        print(f"\n[ERROR] Empty Item name*: {len(issues_found['empty_name'])} rows")
        print(f"   Rows: {issues_found['empty_name'][:20]}{'...' if len(issues_found['empty_name']) > 20 else ''}")
        total_issues += len(issues_found['empty_name'])
    
    if issues_found['empty_base_unit']:
        print(f"\n[ERROR] Empty Base Unit (x): {len(issues_found['empty_base_unit'])} rows")
        print(f"   Rows: {issues_found['empty_base_unit'][:20]}{'...' if len(issues_found['empty_base_unit']) > 20 else ''}")
        total_issues += len(issues_found['empty_base_unit'])
    
    if issues_found['invalid_cost']:
        print(f"\n[WARNING] Invalid cost values: {len(issues_found['invalid_cost'])} rows")
        for row, val in issues_found['invalid_cost'][:10]:
            print(f"   Row {row}: {val}")
        total_issues += len(issues_found['invalid_cost'])
    
    if issues_found['invalid_vat_rate']:
        print(f"\n[WARNING] Invalid VAT rate values: {len(issues_found['invalid_vat_rate'])} rows")
        for row, val in issues_found['invalid_vat_rate'][:10]:
            print(f"   Row {row}: {val}")
        total_issues += len(issues_found['invalid_vat_rate'])
    
    if total_issues == 0:
        print("\n[OK] No obvious data quality issues found in sample rows!")
    else:
        print(f"\n[WARNING] Found {total_issues} potential issues in first 10 rows")
        print("   This might explain why many items are failing validation!")
    
except Exception as e:
    print(f"\nERROR: {e}")
    import traceback
    traceback.print_exc()

print("\n" + "=" * 80)
