"""
Analyze Excel data for common issues that cause validation failures
"""
import os
import sys

try:
    import openpyxl
except ImportError:
    print("Installing openpyxl...")
    os.system(f"{sys.executable} -m pip install openpyxl")
    import openpyxl

excel_path = r"C:\Users\Envy\Downloads\pharmasight_template.xlsx"

print("=" * 80)
print("Excel Data Quality Analysis")
print("=" * 80)

wb = openpyxl.load_workbook(excel_path, data_only=True)
sheet = wb.active

# Get headers
headers = []
for cell in sheet[1]:
    headers.append(str(cell.value).strip() if cell.value else "")

print(f"\nAnalyzing {sheet.max_row - 1} data rows...")

# Analyze all rows
issues = {
    'empty_name': 0,
    'empty_base_unit': 0,
    'none_base_unit': 0,
    'invalid_cost': 0,
    'none_vat_rate': 0,
    'invalid_vat_rate': 0,
    'empty_category': 0,
}

sample_issues = {
    'empty_name': [],
    'empty_base_unit': [],
    'none_base_unit': [],
    'invalid_cost': [],
    'none_vat_rate': [],
    'invalid_vat_rate': [],
}

for row_idx in range(2, sheet.max_row + 1):
    row_data = {}
    for col_idx, header in enumerate(headers, 1):
        cell_value = sheet.cell(row=row_idx, column=col_idx).value
        row_data[header] = cell_value
    
    item_name = str(row_data.get('Item name*', '')).strip() if row_data.get('Item name*') else ''
    base_unit = row_data.get('Base Unit (x)')
    base_unit_str = str(base_unit).strip() if base_unit else ''
    cost = row_data.get('Price_List_Last_Cost') or row_data.get('Price_List_Average_Cost')
    vat_rate = row_data.get('Price_List_Tax_Percentage') or row_data.get('Tax Rate')
    category = row_data.get('Category', '')
    
    # Check issues
    if not item_name or item_name == '' or item_name.lower() == 'none':
        issues['empty_name'] += 1
        if len(sample_issues['empty_name']) < 5:
            sample_issues['empty_name'].append(row_idx)
    
    if base_unit is None:
        issues['none_base_unit'] += 1
        if len(sample_issues['none_base_unit']) < 5:
            sample_issues['none_base_unit'].append(row_idx)
    elif not base_unit_str or base_unit_str == '' or base_unit_str.lower() == 'none':
        issues['empty_base_unit'] += 1
        if len(sample_issues['empty_base_unit']) < 5:
            sample_issues['empty_base_unit'].append(row_idx)
    
    try:
        cost_val = float(cost) if cost is not None else 0
        if cost_val < 0:
            issues['invalid_cost'] += 1
            if len(sample_issues['invalid_cost']) < 5:
                sample_issues['invalid_cost'].append((row_idx, cost_val))
    except (ValueError, TypeError):
        if cost is not None and str(cost).strip() not in ['', 'None', '0', '0.0']:
            issues['invalid_cost'] += 1
            if len(sample_issues['invalid_cost']) < 5:
                sample_issues['invalid_cost'].append((row_idx, cost))
    
    if vat_rate is None:
        issues['none_vat_rate'] += 1
        if len(sample_issues['none_vat_rate']) < 5:
            sample_issues['none_vat_rate'].append(row_idx)
    else:
        try:
            vat_val = float(vat_rate)
            if vat_val < 0 or vat_val > 100:
                issues['invalid_vat_rate'] += 1
                if len(sample_issues['invalid_vat_rate']) < 5:
                    sample_issues['invalid_vat_rate'].append((row_idx, vat_val))
        except (ValueError, TypeError):
            if str(vat_rate).strip() not in ['', 'None', '0', '0.0']:
                issues['invalid_vat_rate'] += 1
                if len(sample_issues['invalid_vat_rate']) < 5:
                    sample_issues['invalid_vat_rate'].append((row_idx, vat_rate))

print("\n" + "=" * 80)
print("ISSUES SUMMARY")
print("=" * 80)

total_issues = sum(issues.values())
print(f"\nTotal rows with issues: {total_issues} out of {sheet.max_row - 1}")
print(f"Success rate: {((sheet.max_row - 1 - total_issues) / (sheet.max_row - 1) * 100):.1f}%")

if issues['empty_name'] > 0:
    print(f"\n[ERROR] Empty Item name*: {issues['empty_name']} rows")
    print(f"   Sample rows: {sample_issues['empty_name']}")

if issues['none_base_unit'] > 0:
    print(f"\n[ERROR] Base Unit (x) is None: {issues['none_base_unit']} rows")
    print(f"   Sample rows: {sample_issues['none_base_unit']}")

if issues['empty_base_unit'] > 0:
    print(f"\n[ERROR] Empty Base Unit (x): {issues['empty_base_unit']} rows")
    print(f"   Sample rows: {sample_issues['empty_base_unit']}")

if issues['invalid_cost'] > 0:
    print(f"\n[WARNING] Invalid cost values: {issues['invalid_cost']} rows")
    for row, val in sample_issues['invalid_cost'][:5]:
        print(f"   Row {row}: {val}")

if issues['none_vat_rate'] > 0:
    print(f"\n[INFO] VAT rate is None (will default to 0): {issues['none_vat_rate']} rows")
    print(f"   Sample rows: {sample_issues['none_vat_rate'][:10]}")

if issues['invalid_vat_rate'] > 0:
    print(f"\n[WARNING] Invalid VAT rate values: {issues['invalid_vat_rate']} rows")
    for row, val in sample_issues['invalid_vat_rate'][:5]:
        print(f"   Row {row}: {val}")

# Check specific patterns
print("\n" + "=" * 80)
print("PATTERN ANALYSIS")
print("=" * 80)

# Check if Base Unit has common values
base_unit_values = {}
for row_idx in range(2, min(100, sheet.max_row + 1)):  # Sample first 100 rows
    base_unit = sheet.cell(row=row_idx, column=headers.index('Base Unit (x)') + 1).value
    if base_unit:
        base_unit_str = str(base_unit).strip().upper()
        base_unit_values[base_unit_str] = base_unit_values.get(base_unit_str, 0) + 1

print(f"\nBase Unit values (first 100 rows):")
for unit, count in sorted(base_unit_values.items(), key=lambda x: x[1], reverse=True)[:10]:
    print(f"   {unit}: {count}")

# Check VAT rate patterns
vat_rate_values = {}
for row_idx in range(2, min(100, sheet.max_row + 1)):
    vat_rate = sheet.cell(row=row_idx, column=headers.index('Price_List_Tax_Percentage') + 1).value
    vat_rate_str = str(vat_rate) if vat_rate is not None else 'None'
    vat_rate_values[vat_rate_str] = vat_rate_values.get(vat_rate_str, 0) + 1

print(f"\nVAT Rate values (first 100 rows):")
for rate, count in sorted(vat_rate_values.items(), key=lambda x: x[1], reverse=True)[:10]:
    print(f"   {rate}: {count}")

print("\n" + "=" * 80)
